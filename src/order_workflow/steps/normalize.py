"""Step 1 - normalize the incoming document (executor: code + OCR).

Whatever arrives (mail, PDF, Excel, CSV, plain text) is turned into one
NormalizedDocument: readable text plus zero or more tables. No model is
involved; OCR is only a fallback for scanned PDFs.
"""

from __future__ import annotations

import csv
import io
from email import policy
from email.parser import BytesParser
from pathlib import Path

from ..models import ExceptionCode, NormalizedDocument, SourceType, Table

EXTENSION_MAP = {
    ".eml": SourceType.EMAIL,
    ".pdf": SourceType.PDF,
    ".xlsx": SourceType.EXCEL,
    ".xlsm": SourceType.EXCEL,
    ".csv": SourceType.CSV,
    ".txt": SourceType.TEXT,
}

# A row is considered a header row when it contains at least two of these.
HEADER_HINTS = {
    "codice", "cod", "articolo", "sku", "code", "item",
    "descrizione", "desc", "materiale", "prodotto", "description", "material", "product",
    "qta", "qtà", "quantita", "quantità", "qty", "quantity", "tons",
    "um", "u.m", "unit", "uom",
    "prezzo", "price", "€/t", "eur/t",
    "consegna", "delivery", "note", "notes",
}


def run(path: Path | str) -> NormalizedDocument:
    path = Path(path)
    source_type = EXTENSION_MAP.get(path.suffix.lower())
    if source_type is None:
        raise ValueError(f"Unsupported file type: {path.suffix} ({path.name})")
    if source_type == SourceType.EMAIL:
        return _normalize_email(path)
    if source_type == SourceType.PDF:
        return _normalize_pdf(path)
    if source_type == SourceType.EXCEL:
        return _normalize_excel(path)
    if source_type == SourceType.CSV:
        return _normalize_csv(path)
    return _normalize_text(path)


# ------------------------------------------------------------------ email


def _normalize_email(path: Path) -> NormalizedDocument:
    with open(path, "rb") as f:
        msg = BytesParser(policy=policy.default).parse(f)
    doc = NormalizedDocument(source_file=path.name, source_type=SourceType.EMAIL)
    doc.metadata["subject"] = msg.get("Subject", "")
    doc.metadata["from"] = msg.get("From", "")
    doc.metadata["date"] = msg.get("Date", "")

    body = msg.get_body(preferencelist=("plain", "html"))
    text = body.get_content() if body else ""
    if body is not None and body.get_content_type() == "text/html":
        text = _strip_html(text)
    doc.text = text.strip()

    # Attachments: normalize each supported one and merge its content.
    for part in msg.iter_attachments():
        filename = part.get_filename() or ""
        suffix = Path(filename).suffix.lower()
        if suffix not in EXTENSION_MAP:
            doc.warnings.append(f"Attachment ignored (unsupported type): {filename}")
            continue
        payload = part.get_payload(decode=True) or b""
        import tempfile

        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(payload)
            tmp_path = Path(tmp.name)
        try:
            sub = run(tmp_path)
            doc.tables.extend(sub.tables)
            if sub.text:
                doc.text += f"\n\n--- Attachment: {filename} ---\n{sub.text}"
            doc.warnings.extend(sub.warnings)
            doc.metadata[f"attachment:{filename}"] = sub.source_type.value
            doc.ocr_used = doc.ocr_used or sub.ocr_used
        finally:
            tmp_path.unlink(missing_ok=True)
    return doc


def _strip_html(html: str) -> str:
    import re

    text = re.sub(r"<(br|/p|/tr|/div)[^>]*>", "\n", html, flags=re.IGNORECASE)
    text = re.sub(r"<(/td|/th)[^>]*>", "\t", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    return io.StringIO(text).getvalue()


# -------------------------------------------------------------------- pdf


def _normalize_pdf(path: Path) -> NormalizedDocument:
    from pypdf import PdfReader

    doc = NormalizedDocument(source_file=path.name, source_type=SourceType.PDF)
    reader = PdfReader(str(path))
    doc.metadata["pages"] = str(len(reader.pages))
    pages = [page.extract_text() or "" for page in reader.pages]
    doc.text = "\n".join(pages).strip()

    if len(doc.text) < 40:  # likely a scan: try OCR, degrade gracefully
        ocr_text = _try_ocr(path)
        if ocr_text is not None:
            doc.text = ocr_text.strip()
            doc.ocr_used = True
        else:
            doc.warnings.append(
                f"{ExceptionCode.OCR_UNAVAILABLE.value}: PDF has little extractable text and "
                "OCR dependencies (pytesseract/pdf2image) are not installed."
            )
    return doc


def _try_ocr(path: Path) -> str | None:
    try:
        import pdf2image
        import pytesseract
    except ImportError:
        return None
    try:
        images = pdf2image.convert_from_path(str(path))
        return "\n".join(pytesseract.image_to_string(img, lang="ita+eng") for img in images)
    except Exception:
        return None


# ------------------------------------------------------------------ excel


def _normalize_excel(path: Path) -> NormalizedDocument:
    from openpyxl import load_workbook

    doc = NormalizedDocument(source_file=path.name, source_type=SourceType.EXCEL)
    wb = load_workbook(str(path), read_only=True, data_only=True)
    preamble: list[str] = []
    for ws in wb.worksheets:
        rows = [
            ["" if cell is None else str(cell) for cell in row]
            for row in ws.iter_rows(values_only=True)
        ]
        rows = [row for row in rows if any(cell.strip() for cell in row)]
        if not rows:
            continue
        header_idx = _find_header_row(rows)
        if header_idx is None:
            preamble.extend(" ".join(cell for cell in row if cell) for row in rows)
            continue
        preamble.extend(
            " ".join(cell for cell in row if cell) for row in rows[:header_idx]
        )
        table = Table(name=ws.title, headers=rows[header_idx], rows=rows[header_idx + 1 :])
        doc.tables.append(table)
    wb.close()
    doc.text = "\n".join(preamble).strip()
    return doc


def _find_header_row(rows: list[list[str]]) -> int | None:
    for idx, row in enumerate(rows[:10]):
        hits = sum(1 for cell in row if cell.strip().lower().rstrip(".:") in HEADER_HINTS)
        if hits >= 2:
            return idx
    return None


# -------------------------------------------------------------------- csv


def _normalize_csv(path: Path) -> NormalizedDocument:
    doc = NormalizedDocument(source_file=path.name, source_type=SourceType.CSV)
    raw = path.read_text(encoding="utf-8-sig")
    # Frequency-based delimiter detection: Sniffer is unreliable when the
    # file starts with free-text preamble lines (common in real orders).
    delimiter = max(",;\t", key=lambda d: raw.count(d))
    rows = [
        row for row in csv.reader(io.StringIO(raw), delimiter=delimiter)
        if any(c.strip() for c in row)
    ]
    if not rows:
        doc.warnings.append("CSV file is empty.")
        return doc
    header_idx = _find_header_row(rows)
    if header_idx is None:
        header_idx = 0
    doc.text = "\n".join(
        " ".join(cell for cell in row if cell) for row in rows[:header_idx]
    ).strip()
    doc.tables.append(Table(name=path.stem, headers=rows[header_idx], rows=rows[header_idx + 1 :]))
    return doc


# ------------------------------------------------------------------- text


def _normalize_text(path: Path) -> NormalizedDocument:
    return NormalizedDocument(
        source_file=path.name,
        source_type=SourceType.TEXT,
        text=path.read_text(encoding="utf-8").strip(),
    )
